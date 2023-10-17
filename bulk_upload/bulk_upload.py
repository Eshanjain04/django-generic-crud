"""
API -> (name,document_id)
     document = DocumentModel(document_id)   #query to document model
     file = document.file     #retrieve an excel file

     bulk_upload_service(name,file) ->
            model = Model(name)
            if model exists
                   get excel headers in excel_headers
                   get headers from model.headers
                   validate_headers(excel_headers,model.headers)       #function to validate headers
                   if validate_headers is true
                          get excel_data row wise
                          validate data from model
                          if validation_data(excel_data,model_validation)         #function to validate data
                              model.model.create()
                         return validation error
                   return error for headers
            return model does not exists header
"""
from vehicle.models import ExcelData
from openpyxl import load_workbook
import cerberus
from django.apps import apps


def validate_data(data_schema, data):
    validation_schema = cerberus.Validator(data_schema)
    is_valid = validation_schema.validate(data)
    if is_valid:
        return True, []
    else:
        validation_errors = validation_schema.errors
        return False, validation_errors


def validate_headers(model_headers, excel_headers):
    return set(model_headers) == set(excel_headers)


def get_model_to_be_updated(app_name, model_name):
    try:
        model = apps.get_model(app_label=app_name, model_name=model_name)
        return model
    except LookupError:
        return None


def add_row_data(excel_data, header_key_dict):
    response_dict = {}
    for key, value in excel_data.items():
        response_dict[header_key_dict.get(key)] = value
    return response_dict


def bulk_upload(name, excel_file):
    try:
        model = ExcelData.objects.filter(name=name).first()
        if model:
            model_headers = []
            validation_schema = {}
            header_key_dict = {}

            model_name_to_be_updated = model.model_name
            app_name_to_be_updated = model.app_name
            model_to_be_updated = get_model_to_be_updated(app_name_to_be_updated, model_name_to_be_updated)
            if not model_to_be_updated:
                return False, [
                    f"Model with '{app_name_to_be_updated}_{model_name_to_be_updated}' not found"]

            for item in model.structure:
                model_headers.append(item.get('name'))
                validation_schema[item.get('key')] = item.get('validation')
                header_key_dict[item.get('name')] = item.get('key')

            wb = load_workbook(excel_file)
            sheet = wb.active

            excel_headers = [cell.value for cell in sheet[1]]
            validated_headers = validate_headers(model_headers=model_headers, excel_headers=excel_headers)
            data_to_be_uploaded = []
            if validated_headers:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    excel_data = dict(zip(excel_headers, row))
                    final_data = add_row_data(excel_data=excel_data,
                                              header_key_dict=header_key_dict)
                    is_valid, errors = validate_data(validation_schema, final_data)
                    if is_valid:
                        data_to_be_uploaded.append(model_to_be_updated(**final_data))
                    else:
                        # todo cerebrus validator giving non specific error
                        return False, errors
                try:
                    model_to_be_updated.objects.bulk_create(data_to_be_uploaded)
                except Exception as err:
                    # todo need to handle validation in right way
                    return False, [str(err)]

                return True, None
            else:
                return False, ["Excel file headers do not match the expected headers in the database."]
        else:
            return False, [
                f"Bulk Upload Data with name '{name}' not found"]
    except Exception as err:
        return False, [str(err)]
